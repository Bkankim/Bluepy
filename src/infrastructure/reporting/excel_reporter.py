"""Excel 보고서 생성

openpyxl을 사용하여 스캔 결과를 Excel 파일로 생성합니다.

주요 기능:
- 요약 시트: 서버 정보, 전체 점수, 통계
- 상세 시트: 모든 점검 항목 결과
- 통계 시트: 카테고리별 분포
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ...core.scanner.base_scanner import ScanResult
from ...core.domain.models import Status


class ExcelReporter:
    """Excel 보고서 생성 클래스

    ScanResult를 받아서 Excel 파일을 생성합니다.
    """

    # 상태별 색상 (RGB)
    COLOR_PASS = "C8E6C9"  # 연한 녹색
    COLOR_FAIL = "FFCDD2"  # 연한 빨간색
    COLOR_MANUAL = "FFE082"  # 연한 노란색
    COLOR_HEADER = "2196F3"  # 파란색

    def __init__(self):
        """초기화"""
        self.wb = None

    def generate(
        self, scan_result: ScanResult, output_path: str, server_name: Optional[str] = None
    ) -> str:
        """보고서 생성

        Args:
            scan_result: 스캔 결과
            output_path: 출력 파일 경로
            server_name: 서버 이름 (선택)

        Returns:
            생성된 파일 경로
        """
        # Workbook 생성
        self.wb = Workbook()

        # 기본 시트 제거
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # 시트 생성
        self._create_summary_sheet(scan_result, server_name)
        self._create_detail_sheet(scan_result)
        self._create_statistics_sheet(scan_result)

        # 파일 저장
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_file)

        return str(output_file)

    def _create_summary_sheet(self, scan_result: ScanResult, server_name: Optional[str]):
        """요약 시트 생성

        Args:
            scan_result: 스캔 결과
            server_name: 서버 이름
        """
        ws = self.wb.create_sheet("요약", 0)

        # 헤더
        ws["A1"] = "BluePy 2.0 보안 점검 보고서"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # 기본 정보
        row = 3
        ws[f"A{row}"] = "서버 정보"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")

        row += 1
        ws[f"A{row}"] = "서버 ID:"
        ws[f"B{row}"] = scan_result.server_id
        row += 1
        ws[f"A{row}"] = "서버 이름:"
        ws[f"B{row}"] = server_name or "-"
        row += 1
        ws[f"A{row}"] = "플랫폼:"
        ws[f"B{row}"] = scan_result.platform
        row += 1
        ws[f"A{row}"] = "스캔 시각:"
        ws[f"B{row}"] = scan_result.scan_time.strftime("%Y-%m-%d %H:%M:%S")

        # 점수
        row += 2
        ws[f"A{row}"] = "전체 점수"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")

        row += 1
        ws[f"A{row}"] = "점수:"
        ws[f"B{row}"] = f"{scan_result.score:.1f} / 100"
        ws[f"B{row}"].font = Font(size=14, bold=True)

        # 통계
        row += 2
        ws[f"A{row}"] = "점검 통계"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")

        row += 1
        ws[f"A{row}"] = "전체 항목:"
        ws[f"B{row}"] = scan_result.total

        row += 1
        ws[f"A{row}"] = "양호 (PASS):"
        ws[f"B{row}"] = scan_result.passed
        ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_PASS, fill_type="solid")

        row += 1
        ws[f"A{row}"] = "취약 (FAIL):"
        ws[f"B{row}"] = scan_result.failed
        ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_FAIL, fill_type="solid")

        row += 1
        ws[f"A{row}"] = "수동 점검 (MANUAL):"
        ws[f"B{row}"] = scan_result.manual
        ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_MANUAL, fill_type="solid")

        # 열 너비 조정
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _create_detail_sheet(self, scan_result: ScanResult):
        """상세 시트 생성

        Args:
            scan_result: 스캔 결과
        """
        ws = self.wb.create_sheet("상세 결과")

        # 헤더
        headers = ["번호", "규칙 ID", "상태", "메시지", "시각"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 데이터
        row = 2
        for idx, (rule_id, check_result) in enumerate(sorted(scan_result.results.items()), start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=rule_id)
            ws.cell(row=row, column=3, value=check_result.status.value)
            ws.cell(row=row, column=4, value=check_result.message)
            ws.cell(row=row, column=5, value=check_result.timestamp.strftime("%H:%M:%S"))

            # 상태별 색상
            status_cell = ws.cell(row=row, column=3)
            if check_result.status == Status.PASS:
                status_cell.fill = PatternFill(start_color=self.COLOR_PASS, fill_type="solid")
            elif check_result.status == Status.FAIL:
                status_cell.fill = PatternFill(start_color=self.COLOR_FAIL, fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color=self.COLOR_MANUAL, fill_type="solid")

            row += 1

        # 열 너비 조정
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 12

        # 자동 필터
        ws.auto_filter.ref = f"A1:E{row-1}"

    def _create_statistics_sheet(self, scan_result: ScanResult):
        """통계 시트 생성

        Args:
            scan_result: 스캔 결과
        """
        ws = self.wb.create_sheet("통계")

        # 헤더
        ws["A1"] = "카테고리별 통계"
        ws["A1"].font = Font(size=14, bold=True)

        # TODO: 실제로는 RuleMetadata에서 카테고리 정보를 가져와야 함
        # 현재는 샘플 데이터
        row = 3
        headers = ["카테고리", "전체", "양호", "취약", "수동"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")

        # 샘플 데이터
        categories = [
            ("계정관리", 15, 12, 2, 1),
            ("파일 및 디렉터리 관리", 20, 15, 3, 2),
            ("서비스 관리", 35, 28, 5, 2),
            ("패치 관리", 1, 0, 0, 1),
            ("로그 관리", 2, 0, 0, 2),
        ]

        row += 1
        for category, total, passed, failed, manual in categories:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=passed)
            ws.cell(row=row, column=4, value=failed)
            ws.cell(row=row, column=5, value=manual)
            row += 1

        # 열 너비 조정
        ws.column_dimensions["A"].width = 30
        for col in ["B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 12


__all__ = [
    "ExcelReporter",
]
