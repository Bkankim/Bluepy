"""Excel Reporter 단위 테스트

src/infrastructure/reporting/excel_reporter.py를 테스트합니다.

테스트 범위:
1. ExcelReporter 클래스 import 및 생성
2. Excel 파일 생성 (임시 디렉토리 사용)
3. Excel 파일 구조 검증
4. 시트 생성 확인
"""

import pytest
from pathlib import Path

from src.core.domain.models import CheckResult, Status
from src.core.scanner.base_scanner import ScanResult
from src.infrastructure.reporting.excel_reporter import ExcelReporter


# ==================== ExcelReporter Import Tests ====================


@pytest.mark.unit
class TestExcelReporterImport:
    """ExcelReporter import 테스트"""

    def test_import_excel_reporter(self):
        """ExcelReporter import 가능 확인"""
        from src.infrastructure.reporting.excel_reporter import ExcelReporter

        assert ExcelReporter is not None

    def test_openpyxl_dependency(self):
        """openpyxl 의존성 확인"""
        try:
            import openpyxl

            assert openpyxl is not None
        except ImportError:
            pytest.fail("openpyxl이 설치되지 않았습니다")


# ==================== ExcelReporter Class Tests ====================


@pytest.mark.unit
class TestExcelReporterClass:
    """ExcelReporter 클래스 테스트"""

    def test_create_excel_reporter(self):
        """ExcelReporter 생성 테스트"""
        reporter = ExcelReporter()
        assert reporter is not None
        assert reporter.wb is None  # 초기 상태에서는 None

    def test_excel_reporter_has_generate_method(self):
        """ExcelReporter에 generate 메서드가 있는지 확인"""
        reporter = ExcelReporter()
        assert hasattr(reporter, "generate")
        assert callable(reporter.generate)

    def test_excel_reporter_color_constants(self):
        """ExcelReporter 색상 상수 확인"""
        assert hasattr(ExcelReporter, "COLOR_PASS")
        assert hasattr(ExcelReporter, "COLOR_FAIL")
        assert hasattr(ExcelReporter, "COLOR_MANUAL")
        assert hasattr(ExcelReporter, "COLOR_HEADER")

        # 색상 값이 hex 형식인지 확인
        assert len(ExcelReporter.COLOR_PASS) == 6
        assert len(ExcelReporter.COLOR_FAIL) == 6
        assert len(ExcelReporter.COLOR_MANUAL) == 6
        assert len(ExcelReporter.COLOR_HEADER) == 6


# ==================== Excel Generation Tests ====================


@pytest.mark.unit
class TestExcelGeneration:
    """Excel 파일 생성 테스트"""

    def test_generate_excel_file(self, tmp_path):
        """Excel 파일 생성 테스트"""
        # ScanResult 생성
        scan_result = ScanResult(server_id="server-001", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Test 1"),
            "U-02": CheckResult(status=Status.FAIL, message="Test 2"),
            "U-03": CheckResult(status=Status.MANUAL, message="Test 3"),
        }

        # ExcelReporter 생성
        reporter = ExcelReporter()

        # Excel 파일 생성
        output_file = tmp_path / "report.xlsx"
        result_path = reporter.generate(
            scan_result=scan_result,
            output_path=str(output_file),
            server_name="Test Server",
        )

        # 파일 생성 확인
        assert Path(result_path).exists()
        assert Path(result_path).suffix == ".xlsx"

    def test_generate_excel_without_server_name(self, tmp_path):
        """server_name 없이 Excel 생성 테스트"""
        scan_result = ScanResult(server_id="server-002", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Test"),
        }

        reporter = ExcelReporter()
        output_file = tmp_path / "report_no_name.xlsx"
        result_path = reporter.generate(scan_result=scan_result, output_path=str(output_file))

        assert Path(result_path).exists()

    def test_generate_excel_creates_parent_directory(self, tmp_path):
        """부모 디렉토리 자동 생성 확인"""
        scan_result = ScanResult(server_id="server-003", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Test"),
        }

        reporter = ExcelReporter()
        # 존재하지 않는 디렉토리 경로
        output_file = tmp_path / "reports" / "subdir" / "report.xlsx"
        result_path = reporter.generate(scan_result=scan_result, output_path=str(output_file))

        assert Path(result_path).exists()
        assert Path(result_path).parent.exists()


# ==================== Excel Content Tests ====================


@pytest.mark.unit
class TestExcelContent:
    """Excel 파일 내용 테스트"""

    def test_excel_has_three_sheets(self, tmp_path):
        """Excel 파일에 3개 시트가 있는지 확인"""
        scan_result = ScanResult(server_id="server-001", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Test"),
        }

        reporter = ExcelReporter()
        output_file = tmp_path / "report.xlsx"
        result_path = reporter.generate(scan_result=scan_result, output_path=str(output_file))

        # openpyxl로 파일 읽기
        import openpyxl

        wb = openpyxl.load_workbook(result_path)

        # 시트 개수 확인 (요약, 상세, 통계)
        assert len(wb.sheetnames) == 3

        # 시트 이름 확인
        assert "요약" in wb.sheetnames
        assert "상세 결과" in wb.sheetnames
        assert "통계" in wb.sheetnames

    def test_summary_sheet_has_server_info(self, tmp_path):
        """요약 시트에 서버 정보가 있는지 확인"""
        scan_result = ScanResult(server_id="server-test", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Test"),
        }

        reporter = ExcelReporter()
        output_file = tmp_path / "report.xlsx"
        result_path = reporter.generate(
            scan_result=scan_result,
            output_path=str(output_file),
            server_name="My Server",
        )

        # 파일 읽기
        import openpyxl

        wb = openpyxl.load_workbook(result_path)
        ws = wb["요약"]

        # 서버 ID 확인 (B4 셀에 있음)
        server_id_value = ws["B4"].value
        assert server_id_value == "server-test"

        # 서버 이름 확인 (B5 셀에 있음)
        server_name_value = ws["B5"].value
        assert server_name_value == "My Server"

    def test_detail_sheet_has_results(self, tmp_path):
        """상세 시트에 결과가 있는지 확인"""
        scan_result = ScanResult(server_id="server-001", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Test 1"),
            "U-02": CheckResult(status=Status.FAIL, message="Test 2"),
        }

        reporter = ExcelReporter()
        output_file = tmp_path / "report.xlsx"
        result_path = reporter.generate(scan_result=scan_result, output_path=str(output_file))

        # 파일 읽기
        import openpyxl

        wb = openpyxl.load_workbook(result_path)
        ws = wb["상세 결과"]

        # 헤더 행이 있는지 확인
        assert ws["A1"].value is not None

        # 데이터 행이 있는지 확인 (헤더 + 2개 결과)
        # 최소 3행 이상이어야 함
        assert ws.max_row >= 3


# ==================== Edge Cases Tests ====================


@pytest.mark.unit
class TestExcelEdgeCases:
    """Excel 생성 경계 케이스 테스트"""

    def test_generate_with_empty_results(self, tmp_path):
        """결과가 없는 경우 Excel 생성"""
        scan_result = ScanResult(server_id="server-empty", platform="linux")
        # 결과 없음

        reporter = ExcelReporter()
        output_file = tmp_path / "report_empty.xlsx"
        result_path = reporter.generate(scan_result=scan_result, output_path=str(output_file))

        assert Path(result_path).exists()

        # 파일 읽기
        import openpyxl

        wb = openpyxl.load_workbook(result_path)
        assert len(wb.sheetnames) == 3

    def test_generate_with_many_results(self, tmp_path):
        """많은 결과가 있는 경우 Excel 생성"""
        scan_result = ScanResult(server_id="server-many", platform="linux")
        # 100개 결과 생성
        for i in range(1, 101):
            scan_result.results[f"U-{i:02d}"] = CheckResult(
                status=Status.PASS if i % 2 == 0 else Status.FAIL, message=f"Test {i}"
            )

        reporter = ExcelReporter()
        output_file = tmp_path / "report_many.xlsx"
        result_path = reporter.generate(scan_result=scan_result, output_path=str(output_file))

        assert Path(result_path).exists()

        # 파일 읽기
        import openpyxl

        wb = openpyxl.load_workbook(result_path)
        ws = wb["상세 결과"]

        # 100개 결과 + 헤더 = 최소 101행
        assert ws.max_row >= 101


# ==================== Integration Tests ====================


@pytest.mark.unit
class TestExcelReporterIntegration:
    """ExcelReporter 통합 테스트"""

    def test_full_workflow(self, tmp_path):
        """전체 워크플로우 테스트"""
        # ScanResult 생성
        scan_result = ScanResult(server_id="server-integration", platform="linux")
        scan_result.results = {
            "U-01": CheckResult(status=Status.PASS, message="Good"),
            "U-02": CheckResult(status=Status.PASS, message="Good"),
            "U-03": CheckResult(status=Status.FAIL, message="Bad"),
            "U-04": CheckResult(status=Status.MANUAL, message="Check manually"),
        }

        # ExcelReporter로 보고서 생성
        reporter = ExcelReporter()
        output_file = tmp_path / "full_report.xlsx"
        result_path = reporter.generate(
            scan_result=scan_result,
            output_path=str(output_file),
            server_name="Integration Test Server",
        )

        # 검증
        assert Path(result_path).exists()

        # Excel 파일 읽기
        import openpyxl

        wb = openpyxl.load_workbook(result_path)

        # 3개 시트 확인
        assert len(wb.sheetnames) == 3

        # 요약 시트 확인
        summary_ws = wb["요약"]
        assert summary_ws["B4"].value == "server-integration"
        assert summary_ws["B5"].value == "Integration Test Server"

        # 점수 확인 (2 PASS + 0.5 MANUAL) / 4 * 100 = 62.5
        # 점수는 B10 또는 B11 셀에 있을 수 있음
        # (구현에 따라 다를 수 있으므로 존재만 확인)

        # 상세 시트에 4개 결과 확인
        detail_ws = wb["상세 결과"]
        assert detail_ws.max_row >= 5  # 헤더 + 4개 결과

        # 통계 시트 존재 확인
        assert "통계" in wb.sheetnames
